#! python3
#
# version
# -------------------------------------------------------------------------------------------------------
# 1.0 A useful project to learn Python and Git, GitHub
#
# 2.0 Simplify .lnx file only ith .htm, .html entries and handle timeouts more sensibly
#
# TODO version 2.0+ extract urls from any text based file
#

# -------------------------------------------------------------------------------------------------------
# xLinks.py - add each link in given file to specified excel workbook/sheet
#
#  A .lnx file contains csv formatted lines of:
#
#           <path>/file.htm or <path>/file.html
#
#    each containing one or more  <a href="https://www.w3schools.com">Visit W3Schools.com!</a>  links to process
#
# If the workbook/sheet specified does not exist create it... Enforce/check the format of the
# sheet (e.g. some magic number in a cell).
#
# If a link cannot be succesfully processed e.g. timeout the link is written to output file 'unprocessed.lnx' in
# same path as input .lnx file
#
# -------------------------------------------------------------------------------------------------------------
# TODO display raw content from url (for debug/interest)
# TODO decode 200, 404 etc.
# TODO ignore UTF-8 decode erors
# TODO unquote quoted urls , no %20 %3F
# -------------------------------------------------------------------------------------------------------------

import sys
import os
import csv
import requests
import bs4
import lxml
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Color
import datetime
import time
import logging

usage = '''
- Usage:
-        py.exe xllinks.py <links filename> <target xlsx> <sheet name>
-
-           <links filename> must end with '.lnx'
'''

# UTF-8, Unicode etc. problems...
# ----------------------------------------------------------------------------------------------------
# From https://stackoverflow.com/questions/36833357/python-correct-encoding-of-website-beautiful-soup
#
# It's not BeautifulSoup's fault. You can see this by printing out encodedText, before you
# ever use BeautifulSoup: the non-ASCII characters are already gibberish.
#
# The problem here is that you are mixing up bytes and characters. For a good overview of
# the difference, read one of Joel's articles, but the gist is that bytes are, well,
# bytes (groups of 8 bits without any further meaning attached), whereas characters are
# the things that make up strings of text. ENCODING TURNS CHARACTERS INTO BYTES, AND
# # DECODING TURNS BYTES BACK INTO CHARACTERS.
#
# A look at the requests documentation shows that r.text is made of characters, not
# bytes. You shouldn't be encoding it. If you try to do so, you will make a byte string,
# and when you try to treat that as characters, bad things will happen.
#
# There are two ways to get around this:
#
# Use the raw undecoded bytes, which are stored in r.content, as Martijn suggested.
# Then you can decode them yourself to turn them into characters.
#
# Let requests do the decoding, but just make sure it uses the right codec. Since you
# know that's UTF-8 in this case, you can set r.encoding = 'utf-8'. If you do this
# before you access r.text, then when you do access r.text, it will have been properly
# decoded, and you get a character string. You don't need to mess with character
# encodings at all.
#
# Incidentally, Python 3 makes it somewhat easier to maintain the difference between
# character strings and byte strings, because it requires you to use different types
# of objects to represent them.
#

# ---bs4_parser = 'html.parser'
bs4_parser = 'lxml'

MAGICVALUE = "#01#01#01#"


class xlHandle:

    def __init__(self, wbPath, wsName):
        self.wbPath = wbPath
        self.wsName = wsName
        self.wb = None
        self.ws = None
        self.columnData = {}
        self.headerRow = True


    def setupWorkbook(self):
        self.wb, self.ws = setupWorkbook(self, self.wbPath, self.wsName)


    def recordColumnMaxWidth(self, colNo, colWidth): # for given column set/reset maximum width
        col = get_column_letter(colNo)
        self.columnData.setdefault(col, {})
        self.columnData[col].setdefault('maxWidth' , 0)
        currentMaxWidth =  self.columnData[col]['maxWidth']
        if colWidth > currentMaxWidth:
            self.columnData[col]['maxWidth'] = colWidth


    def scanColumnWidths(self): # scan sheet for max data width
        for r in range(1, self.ws.max_row + 1):
            for c in range (1, self.ws.max_column + 1):
                dataLength = len(str(self.ws.cell(row=r, column=c).value))
                self.recordColumnMaxWidth(c, dataLength)


    def setColumnWidths(self):
        for c in range(1, self.ws.max_column + 1):
            col = get_column_letter(c)
            self.ws.column_dimensions[col].width = self.columnData[col]['maxWidth'] * 1.5  # --- fudge factor ---


    def setFontAndColour(self):

        cellFont = Font(name='Times New Roman', size=12)
        cellFontBold = Font(name='Times New Roman', size=12, bold=True)
        headerFillcolor = Color(rgb='00C4D79B')
        oddFillcolor = Color(rgb='00E9E17F')
        cellHeaderFill = PatternFill(patternType='solid', fgColor=headerFillcolor)
        cellOddFill = PatternFill(patternType='solid', fgColor=oddFillcolor)

        for r in range(1, self.ws.max_row + 1):
            for c in range(1, self.ws.max_column + 1):

                # 1) font , size
                if self.headerRow and r == 1:
                    self.ws.cell(row=r, column=c).font = cellFontBold
                else:
                    self.ws.cell(row=r, column=c).font = cellFont

                # 2) color :  header row b/g colour , odd row b/g colour
                if r == 1:
                    if self.headerRow:
                        self.ws.cell(row=r, column=c).fill = cellHeaderFill
                    else:
                        self.ws.cell(row=r, column=c).fill = cellOddFill
                else:
                    if r & 1:
                        self.ws.cell(row=r, column=c).fill = cellOddFill



    def adjustColumnWidths(self): # adjust the column widths to reflect data
        if self.ws is not None:
            self.setFontAndColour()
            self.scanColumnWidths()
            self.setColumnWidths()


    def finishWorkbook(self):
        if self.wb is not None:
            self.adjustColumnWidths()


    def saveWorkbook(self):
        if self.wb is not None:
            self.finishWorkbook()
            self.wb.save(self.wbPath)

    def closeWorkbook(self): # same as saveWorkbook plus...?
        if self.wb is not None:
            self.finishWorkbook()
            self.wb.save(self.wbPath)


def printUsage():
    print(usage)

def setupWorkbook(xlH, file, sheet):
    # open file if it exists and chekc if writable, else create new workbook / sheet
    # check if sheet exists else, add it to workbook
    wb = None
    ws = None
    try:
        # does workbook exist? Create if not...
        try:
            wb = openpyxl.load_workbook(file)

            # check if writeable...
            try:
                wb.save(file)
            except Exception as e:
                logging.critical('Exception [{0}][{1}]'.format(e, file))
                raise PermissionError # pass it on...

        except PermissionError as e: # pass it on...
            raise PermissionError

        except Exception as e:

            logging.info('---exception [{0}][{1}][{2}]---'.format(e, file, sheet))
            
            # assume workbook does not exist , create new workbook...

            wb = openpyxl.Workbook()
            # add required sheet (replace sheet 0)...
            ws = wb.worksheets[0]
            ws.title = sheet
            setupWorksheet(xlH, ws)
            wb.save(file)
            wb = openpyxl.load_workbook(file)

        # does sheet exist, if not create it
        if not sheet in wb.sheetnames:
            wb.create_sheet(sheet)
            ws = wb[sheet]
            setupWorksheet(xlH, ws)

        ws = wb[sheet]
        if not validWorksheet(ws):
            logging.critical('Not MAGIC!!! [{0}][{1}]'.format(file, sheet))
            wb.close()
            wb = None
            ws = None

    except Exception as e:
        logging.error('Exception [{0}][{1}][{2}]'.format(e, file, sheet))
        wb = None
        ws = None

    return wb, ws


def validWorksheet(ws):  # application specific worksheet code
    magic = ws['A1'].value
    return magic == MAGICVALUE


def setupWorksheet(xlH, ws):  # application specific worksheet code
    ws['A1'] = MAGICVALUE
    ws['B1'] = 'Group'
    ws['C1'] = 'Result code'
    ws['D1'] = 'Reason'
    ws['E1'] = 'Link'
    ws['F1'] = 'Title'
    ws['G1'] = 'Date entered'
    ws.freeze_panes = ws['A2']

    logging.info('setup worksheet:max row={0}'.format(ws.max_row))


def addLink(xlH, theGroup, theLink, theTitle, theResultCode, theResultReason):  # application specific worksheet code

    ws = xlH.ws

    # find next available row...
    lastRow = ws.max_row
    nextRow = lastRow + 1

    ws['A' + str(nextRow)] = MAGICVALUE
    ws['B' + str(nextRow)] = theGroup
    ws['C' + str(nextRow)] = theResultCode
    ws['D' + str(nextRow)] = theResultReason
    ws['E' + str(nextRow)] = theLink
    ws['F' + str(nextRow)] = theTitle
    now = datetime.datetime.now()
    ws['G' + str(nextRow)] = now.strftime("%Y-%m-%d-%H-%M-%S")

def addToDump(unprocessed, line):
    upf = open(unprocessed,"a")
    upf.write(line+'\n')
    upf.close()

def stamp(stampfile, message):
    sf = open(stampfile,"w")
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y-%m-%d-%H-%M-%S") + ' ' + message + '\n'
    sf.write(timestamp)
    sf.close()

def checkUrl(url):

    # check url and catch Request exceptions , convert to 'error' string

    timeoutSeconds = 10

    start_time = got_time = end_time = 0

    start_time = time.perf_counter()

    rescode = 0
    reason = '?'
    title = '?'

    try:

        req_headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64; rv:77.0) Gecko/20100101 Firefox/77.0'
        }

        res = requests.get(url, headers=req_headers, timeout=timeoutSeconds)

        got_time = time.perf_counter()

        rescode = res.status_code
        reason = res.reason
        res.raise_for_status()
    # TODO ... ? text = res.text.encode('UTF-8', ignore errors ...) ?
    #          ? OR text = res.content.decode('UTF-8' , ignore errors... )?
        slop = bs4.BeautifulSoup(res.text, bs4_parser)
        # --- TODO --- UNICODE/UTF-8 invalid characters in HTML
        # --- ? slop = bs4.BeautifulSoup(res.read().decode('utf-8', 'ignore'),bs4_parser)
        title = slop.find('title').get_text(strip=True)

    except TimeoutError as e:
        reason = 'Time out exception [{0}][{1}][timeout = {2}s]'.format(e, url, timeoutSeconds)
        logging.info(reason)
        rescode = 0
        title = '- TIME OUT -'

    except requests.exceptions.RequestException as e:
        reason = 'Request Exception [{0}][{1}]'.format(e, url)
        logging.info(reason)
        if rescode == 0: # Request exception with no HTTP status code set...
            title = '-? POSSIBLE TIME OUT retry count exceeded ?-'

    except Exception as e:
        reason = 'Exception [{0}][{1}]'.format(e, url)
        logging.info(reason)
        rescode = 0
        title = '- PROBLEM -'

    end_time = time.perf_counter()

    if got_time:
        time0 = got_time - start_time
        time1 = end_time - start_time
    else:
        time0 = 0
        time1 = end_time - start_time

    logging.debug('>>>checkURL took {0:10.5f} / {1:10.5f} seconds for url[{2}]-[{3}]'.format(time0, time1,  url, rescode))
    logging.debug('<<< rescode [{0}] , reason [{1}] , url [{2}] , title [{3}]'.format(rescode, reason, url, title))
    return (rescode, reason, title)

def nextEmbeddedLink(soup):
    # find all anchor tags...
    #
    # [1] href_tags = soup.find_all(href=True)
    # OR
    # [2] for a in soup.find_all('a', href=True):
    #       print "Found the URL:", a['href']
    #
    for a in soup.find_all('a'):
        href = a.get('href')
        if href.startswith('http://') or href.startswith('https://'):
            yield href

def processHTML(xlH, htmlFile, unprocessed, stampfile):

    # given HTML file name (hopefully)... TODO check if html, UTF-8  ?
    if os.path.isfile(htmlFile):
        logging.info('Processing HTML FILE [{0}]...'.format(htmlFile))
        # ??? errors='ignore', 'replace', 'backslashreplace'???
        fp = open(htmlFile, encoding='utf-8', errors='backslashreplace')
        soup = bs4.BeautifulSoup(fp, bs4_parser)
        for link in nextEmbeddedLink(soup):
            processLink(xlH, link, unprocessed, stampfile)
        fp.close()
    else:
        print('Processing HTML FILE [{0}] -- NO SUCH FILE --'.format(htmlFile))

def processLink(xlH, link, unprocessed, stampfile, _static={'counter': 0}):

    _static['counter'] += 1
    modval1000 = _static['counter'] % 1000

    rescode, reason, title = checkUrl(link)

    # defend against unknwown events...  occasionally save workbook in case
    if modval1000 == 0:
        xlH.saveWorkbook()

    # Append to the specified workbook/sheet...
    if rescode == 200:
        addLink(xlH, 'UNCLASSIFIED', link, title, rescode, reason)
    else:
        addToDump(unprocessed,'{0} {1}'.format(link,rescode))

    if rescode  != 200:
        logging.info('{3:4d} Result [{1} {4}] Link [{0}] Title[{2}]'.format(link, rescode, title, _static['counter'], reason))

    stamp(stampfile,'{0:4d}'.format(_static['counter']))

    return 0

def processLNXFile(xlH, lnxFile, unprocessed, stampfile):

    if lnxFile.endswith('.lnx') and os.path.isfile(lnxFile):
        logging.info('Processing LNX FILE [{0}]...'.format(lnxFile))
        with open(lnxFile, mode='r') as csv_file:
            csv_reader = csv.DictReader(csv_file, fieldnames=["col0", "col1"], restval='')
            for row in csv_reader:
                col0 = row['col0'].strip()
                if col0.endswith('.htm') or col0.endswith('.html'):
                    processHTML(xlH, col0, unprocessed, stampfile)
                else:
                    logging.info('Processing LNX FILE [{0}] -- unknown entry --'.format(col0))
            csv_file.close()
    else:
        logging.info('Processing LNX FILE [{0}] -- NO SUCH FILE --'.format(lnxFile))


def main():

    # What to do?
    if len(sys.argv) != 4:
        printUsage()
        return -1

    logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

    links = sys.argv[1]
    xlfile = sys.argv[2]
    xlsheet = sys.argv[3]

    unprocessed = os.path.dirname(os.path.abspath(links))+os.path.sep+'unprocessed.lnx'
    stampfile = os.path.dirname(os.path.abspath(links))+os.path.sep+'stamp'

    xlH = xlHandle(xlfile, xlsheet)
    try:
        # set up excel target: workbook and sheet
        xlH.setupWorkbook()

        if links.endswith('.lnx'):

            if os.path.isfile(links):

                if xlH.ws is not None:
                    processLNXFile(xlH, links, unprocessed, stampfile)

            else:
                logging.info('Processing LNX FILE [{0}] -- NO SUCH FILE --'.format(links))

        else:
            printUsage()

    except requests.exceptions.RequestException as e:
        logging.error('!!! Requests exception !!! [{0}]'.format(e))
    except UnicodeError as e:
        logging.error('!!! UNICODE ERROR !!! [{0}]'.format(e))
    except Exception as e:
        logging.error('!!! Unexpected exception !!! [{0}]'.format(e))
    finally:
       xlH.closeWorkbook()

    del xlH

    stamp(stampfile,'DONE')
    
    return 0


if __name__ == '__main__':

    rc = main()
    sys.exit(rc)
